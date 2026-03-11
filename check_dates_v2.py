from openpyxl import load_workbook
import datetime

wb = load_workbook(filename='backend/Sales_Combined.xlsx', read_only=True)
ws = wb.active
dates = []
unparseable = []
for row in ws.iter_rows(min_row=2, values_only=True):
    val = row[3] # Date column index 3
    if val is not None:
        if isinstance(val, datetime.datetime):
            dates.append(val)
        elif isinstance(val, str):
            try:
                dates.append(datetime.datetime.strptime(val, '%d/%m/%Y'))
            except ValueError:
                unparseable.append(val)
        else:
             # Maybe int/float (excel serial date)? openpyxl usually handles this.
             pass

if dates:
    print(f"Dates found: {len(dates)}")
    print(f"Min: {min(dates)}")
    print(f"Max: {max(dates)}")

if unparseable:
    print(f"Unparseable strings found ({len(unparseable)}):")
    print(unparseable[:10])
